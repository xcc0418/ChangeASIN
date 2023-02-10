import json
from tkinter import *
from tkinter import StringVar, filedialog, Label
import datetime
from tkinter import messagebox
import tkinter.font as tkFont
import pymysql
import openpyxl
import global_var
import os
from tkinter import ttk
import shutil
import find_msku


class Create_msku(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.root = master  # 定义内部变量root
        self.xscroll = Scrollbar(self, orient=HORIZONTAL)
        self.yscroll = Scrollbar(self, orient=VERTICAL)
        self.ft = tkFont.Font(family='microsoft yahei', size=10)
        self.msku = StringVar()
        self.num_msku = StringVar()
        self.sku = StringVar()
        self.country = StringVar()
        self.supplier = StringVar()
        self.itemname = StringVar()
        self.creat()

    def creat(self):
        self.labelframe1 = LabelFrame(self, bd=0)
        self.labelframe1.grid(row=0, column=0, columnspan=2, pady=10, stick=W)
        self.Radiolist = IntVar()
        self.Radiolist2 = IntVar()
        Radiobutton(self.labelframe1, text="单个生成", value=1, variable=self.Radiolist, command=self.jindutiao, font=self.ft).grid(row=0, column=0, padx=10, pady=10)
        Radiobutton(self.labelframe1, text="批量导入", value=2, variable=self.Radiolist, command=self.jindutiao1, font=self.ft).grid(row=0, column=1, padx=10, pady=10)
        self.labelframe2 = LabelFrame(self, bd=0)
        Label(self.labelframe2, text='MSKU模板：', font=self.ft).grid(row=5, column=0, stick=E, padx=0, pady=10)
        Entry(self.labelframe2, textvariable=self.msku).grid(row=5, column=1, columnspan=2, stick=W, padx=0, pady=10)
        Label(self.labelframe2, text='注：MSKU模板无需填写日期，长度不超过23').grid(row=6, column=1, stick=W, pady=10)
        Label(self.labelframe2, text='SKU：', font=self.ft).grid(row=1, column=0, stick=E, padx=0, pady=10)
        Entry(self.labelframe2, textvariable=self.sku).grid(row=1, column=1, columnspan=2, stick=W, padx=0, pady=10)
        Label(self.labelframe2, text='数量：', font=self.ft).grid(row=4, column=0, stick=E, padx=0, pady=10)
        Entry(self.labelframe2, textvariable=self.supplier).grid(row=3, column=1, columnspan=2, stick=W, padx=0, pady=10)
        Label(self.labelframe2, text='供应商：', font=self.ft).grid(row=3, column=0, stick=E, padx=0, pady=10)
        Entry(self.labelframe2, textvariable=self.num_msku).grid(row=4, column=1, columnspan=2, stick=W, padx=0, pady=10)
        Label(self.labelframe2, text='国家：', font=self.ft).grid(row=2, column=0, stick=E, padx=0, pady=10)
        self.cmb = ttk.Combobox(self.labelframe2, width=15, textvariable=self.country, state='readonly', takefocus=False, font=self.ft)
        self.cmb.grid(row=2, column=1, pady=5, stick=W)
        self.cmb['values'] = ['美国', '英国', '加拿大', '日本', '德国', '意大利', '西班牙']
        Button(self.labelframe2, text='批量生成', font=self.ft, command=self.get_msku).grid(row=7, column=0, stick=W, pady=10, padx=10)
        self.labelframe3 = LabelFrame(self, bd=0)
        Label(self.labelframe3, text='文件绝对路径：', font=self.ft).grid(row=1, column=0, stick=E, padx=0, pady=10)
        self.e1 = Entry(self.labelframe3, textvariable=self.itemname, width=30, state='readonly', font=self.ft)
        self.e1.grid(row=1, column=1, columnspan=3, stick=E, padx=10, pady=10)
        Button(self.labelframe3, text='浏览', font=self.ft, command=self.find, width=8).grid(row=1, column=4, stick=W, padx=10, pady=10)
        Button(self.labelframe3, text='导入', font=self.ft, command=self.get_msku, width=15).grid(row=3, column=3, stick=W,pady=10, padx=10)
        # self.labelframe2.grid(row=3)

    def find(self):
        filepath = filedialog.askopenfilename()
        self.e1['state'] = 'normal'
        self.e1.delete(0, 'end')
        self.e1.insert(0, filepath)
        self.e1['state'] = 'readonly'

    def jindutiao(self):
        self.labelframe2.grid(row=3)
        self.labelframe3.grid_forget()

    def jindutiao1(self):
        self.labelframe2.grid_forget()
        self.labelframe3.grid(row=3)

    def sql(self):
        self.connection = pymysql.connect(host='3354n8l084.goho.co',  # 数据库地址
                                          port=24824,
                                          user='test_user',  # 数据库用户名
                                          password='a123456',  # 数据库密码
                                          db='storage',  # 数据库名称
                                          charset='utf8',
                                          cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor()

    def sql_close(self):
        self.cursor.close()
        self.connection.close()

    def get_msku(self):
        index = self.Radiolist.get()
        if index == 1:
            ask = messagebox.askokcancel(message='是否批量生成msku')
            if ask:
                try:
                    msku = self.msku.get().strip()
                    num = int(self.num_msku.get())
                    sku = self.sku.get().strip()
                    country = self.country.get().strip()
                    supplier = self.supplier.get().strip()
                    if msku and num and sku and country and len(msku) <= 23 and supplier:
                        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
                        list_msku = []
                        for i in range(1, num+1):
                            msku_new = f'{msku}-{i}-{time_now[2:]}-T'
                            list_msku.append(msku_new)
                        print(list_msku)
                        self.write_excl(list_msku)
                        self.write_sql(list_msku, sku, country, supplier)
                        messagebox.showinfo(message='MSKU生成成功')
                        self.clear()
                        os.startfile('D:/MSKU生成')
                    else:
                        if not msku:
                            messagebox.showinfo(message='请输入MSKU')
                        if not sku:
                            messagebox.showinfo(message='请输入SKU')
                        if not num:
                            messagebox.showinfo(message='请输入要生成的个数')
                        if not country:
                            messagebox.showinfo(message='请输入国家')
                        if not supplier:
                            messagebox.showinfo(message='请输入供应商')
                        if len(msku) > 30:
                            messagebox.showinfo(message=f'MSKU模板长度为{len(msku)}')
                except Exception as e:
                    messagebox.showinfo(message=e)
        else:
            itemname = self.itemname.get()
            if itemname:
                ask = messagebox.showinfo(message=f'是否导入{itemname}这个文件')
                if ask:
                    wb = openpyxl.load_workbook(itemname, data_only=True)
                    wb_sheet = wb.active
                    row1 = wb_sheet.max_row
                    for i in range(row1, 0, -1):
                        cell_value1 = wb_sheet.cell(row=i, column=1).value
                        if cell_value1:
                            row1 = i
                            break
                    list_msku = []
                    for i in range(2, row1+1):
                        sku = wb_sheet.cell(row=i, column=1).value.strip()
                        msku = wb_sheet.cell(row=i, column=2).value.strip()
                        country = wb_sheet.cell(row=i, column=3).value.strip()
                        supplier = wb_sheet.cell(row=i, column=4).value.strip()
                        num = int(wb_sheet.cell(row=i, column=5).value)
                        list_msku_index = []
                        if sku and len(msku) < 23 and country and supplier and num:
                            time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
                            for j in range(1, num + 1):
                                msku_new = f'{msku}-{j}-{time_now[2:]}-T'
                                list_msku_index.append(msku_new)
                                list_msku.append(msku_new)
                            self.write_sql(list_msku_index, sku, country, supplier)
                    self.write_excl(list_msku)
                    messagebox.showinfo(message='MSKU生成成功')
                    os.startfile('D:/MSKU生成')
            else:
                messagebox.showinfo(message='请先导入表格文件路径')

    def write_sql(self, list_msku, sku, country, supplier):
        self.sql()
        for i in list_msku:
            if supplier:
                sql1 = "select * from `amazon_form`.`pre_msku` where `MSKU` = '%s'" % i
                self.cursor.execute(sql1)
                result = self.cursor.fetchall()
                if result:
                    messagebox.showinfo(message=f'{i}已重复')
                else:
                    sql = "insert into `amazon_form`.`pre_msku`(`ASIN`, `MSKU`, `SKU`, `国家`, `供应商`, `状态`)values" \
                          "('NULL', '%s', '%s', '%s', '%s', '未匹配')" % (i, sku, country, supplier)
                    self.cursor.execute(sql)
            else:
                messagebox.showinfo(message=f'请输入{sku}的供应商')
        self.connection.commit()
        self.sql_close()

    def write_excl(self, list_msku):
        self.mkdir()
        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        wb = openpyxl.Workbook()
        wb_sheet = wb.active
        wb_sheet.append(['MSKU'])
        for i in list_msku:
            wb_sheet.append([i])
        wb.save(f'D:/MSKU生成/msku_{time_now}.xlsx')

    def mkdir(self):
        folder = os.path.exists("D:/MSKU生成")
        if not folder:
            os.makedirs("D:/MSKU生成")

    def clear(self, event=None):
        self.sku.set('')
        self.msku.set('')
        self.supplier.set('')
        self.itemname.set('')
        self.num_msku.set('')


class Exchange(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.root = master  # 定义内部变量root
        self.xscroll = Scrollbar(self, orient=HORIZONTAL)
        self.yscroll = Scrollbar(self, orient=VERTICAL)
        self.ft = tkFont.Font(family='microsoft yahei', size=10)
        self.num = StringVar()
        self.sku = StringVar()
        self.itemname = StringVar()
        self.warehouse = StringVar()
        self.msg = StringVar()
        self.country = StringVar()
        self.creat()

    def sql(self):
        self.connection = pymysql.connect(host='3354n8l084.goho.co',  # 数据库地址
                                          port=24824,
                                          user='test_user',  # 数据库用户名
                                          password='a123456',  # 数据库密码
                                          db='storage',  # 数据库名称
                                          charset='utf8',
                                          cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor()

    def sql_close(self):
        self.cursor.close()
        self.connection.close()

    def creat(self):
        Label(self).grid(row=0, stick=W, pady=10)
        self.farm1 = LabelFrame(self, text='批量导入', font=self.ft)
        self.farm1.grid(row=1, column=1, stick=W, columnspan=4)
        self.farm2 = LabelFrame(self, text='单个SKU申请', font=self.ft)
        self.farm2.grid(row=2, column=1, stick=W, columnspan=4)
        Label(self.farm2, text='SKU：', font=self.ft).grid(row=1, column=0, stick=E, padx=0, pady=10)
        Entry(self.farm2, textvariable=self.sku).grid(row=1, column=1, columnspan=2, stick=W, padx=0, pady=10)
        Label(self.farm2, text='数量：', font=self.ft).grid(row=2, column=0, stick=E, padx=0, pady=10)
        Entry(self.farm2, textvariable=self.num).grid(row=2, column=1, columnspan=2, stick=W, padx=0, pady=10)
        Label(self.farm2, text='国家：', font=self.ft).grid(row=3, column=0, stick=E, padx=0, pady=10)
        Entry(self.farm2, textvariable=self.country).grid(row=3, column=1, columnspan=2, stick=W, padx=0, pady=10)
        Button(self.farm2, text='查询FNSKU', font=self.ft, command=self.get_msg).grid(row=1, column=4, stick=W, pady=10,padx=10)
        Button(self.farm2, text='获取FNSKU', font=self.ft, command=self.get_fnsku).grid(row=2, column=4, stick=W, pady=10,padx=10)
        Label(self.farm1, text='文件绝对路径：', font=self.ft).grid(row=1, column=0, stick=E, padx=0, pady=10)
        self.e1 = Entry(self.farm1, textvariable=self.itemname, width=30, state='readonly', font=self.ft)
        self.e1.grid(row=1, column=1, columnspan=3, stick=E, padx=10, pady=10)
        Label(self.farm1, text='换标仓库：', font=self.ft).grid(row=2, column=0, stick=E, padx=0, pady=10)
        self.cmb = ttk.Combobox(self.farm1, width=30, textvariable=self.warehouse, state='readonly', takefocus=False, font=self.ft)
        self.cmb.grid(row=2, column=1, columnspan=3, pady=5, stick=W)
        self.cmb['values'] = ['工厂仓库', '横中路仓库-加拿大', '横中路仓库-日本', '横中路仓库-美国',
                              '横中路仓库-英国', '横中路仓库-德国', '淘汰-横中路成品仓库-加拿大', '淘汰-横中路成品仓库-日本',
                              '淘汰-横中路成品仓库-美国', '淘汰-横中路成品仓库-英国', '百汇办公室']
        Button(self.farm1, text='浏览', font=self.ft, command=self.find, width=8).grid(row=1, column=4, stick=W, padx=10,pady=10)
        Button(self.farm1, text='获取FNSKU', font=self.ft, command=self.get_fnsku).grid(row=3, column=1, stick=W,pady=10, padx=10)
        Button(self.farm1, text='换标调整', font=self.ft, command=self.read_excl, width=15).grid(row=3, column=2, stick=W,pady=10, padx=10)
        Button(self.farm1, text='清空', font=self.ft, command=self.clear).grid(row=3, column=4, stick=W,pady=10,padx=10)
        Label(self, textvariable=self.msg, font=self.ft, bg='green').grid(row=6, column=1, stick=W, pady=10)

    def find(self):
        filepath = filedialog.askopenfilename()
        self.e1['state'] = 'normal'
        self.e1.delete(0, 'end')
        self.e1.insert(0, filepath)
        self.e1['state'] = 'readonly'

    def get_warehouse(self, warehouse):
        dict_warehouse = {'工厂仓库': 2156, '横中路仓库-加拿大': 1489, '横中路仓库-日本': 1490,
                          '横中路仓库-美国': 1461, '横中路仓库-英国': 1488,
                          '横中路仓库-德国': 2382, '淘汰-横中路成品仓库-加拿大': 1476, '淘汰-横中路成品仓库-日本': 1477,
                          '淘汰-横中路成品仓库-美国': 1399, '淘汰-横中路成品仓库-英国': 1478, '百汇办公室': 414}
        wid = dict_warehouse[warehouse]
        return wid

    def get_fnsku(self, event=None):
        filename = self.itemname.get().strip()
        sku = self.sku.get().strip()
        print(sku)
        if filename:
            ask = messagebox.askokcancel(message=f'是否上传{filename}这个文件')
            if ask:
                try:
                    self.sql()
                    wb = openpyxl.load_workbook(filename)
                    wb_sheet = wb.active
                    row1 = wb_sheet.max_row
                    for i in range(row1, 0, -1):
                        cell_value1 = wb_sheet.cell(row=i, column=1).value
                        if cell_value1:
                            row1 = i
                            break
                    ws = openpyxl.Workbook()
                    ws_sheet = ws.active
                    ws_sheet.append(['*SKU', '品名', '原FNSKU', '调整FNSKU', '调整量'])
                    for i in range(2, row1+1):
                        sku_excl = wb_sheet.cell(row=i, column=1).value.strip()
                        num = int(wb_sheet.cell(row=i, column=2).value)
                        country = wb_sheet.cell(row=i, column=3).value.strip()
                        # num_order = self.get_inventory(sku)
                        # if num <= num_order:
                        sql = "select * from `amazon_form`.`pre_msku` where `SKU` = '%s' and `状态` = '未使用' and " \
                              "`国家` = '%s'" % (sku_excl, country)
                        # print(sql)
                        self.cursor.execute(sql)
                        result = self.cursor.fetchall()
                        print(result)
                        if result and num <= len(result):
                            k = 0
                            product_name, list_fnsku = self.get_productname(sku_excl)
                            for j in result:
                                if j['FNSKU'] in list_fnsku:
                                    k += 1
                                    self.change_sql(j['FNSKU'])
                                    ws_sheet.append([sku_excl, product_name, '', j['FNSKU']])
                                    if k == num:
                                        break
                        else:
                            messagebox.showinfo(message=f'{sku_excl}的fnsku未使用个数不足')
                    time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
                    self.connection.commit()
                    self.mkdir()
                    self.sql_close()
                    self.itemname.set('')
                    ws.save(f'D:/换标调整/换标调整-{time_now}.xlsx')
                    messagebox.showinfo(message='换标模板生成成功')
                    os.startfile("D:/换标调整")
                except Exception as e:
                    messagebox.showinfo(message=e)
        if sku:
            sku_num = self.num.get()
            country = self.country.get().strip()
            if sku_num and country:
                sku_num = int(sku_num)
                print(sku_num)
                ask = messagebox.askokcancel(message=f'是否获取{sku}对应的FNSKU')
                if ask:
                    self.sql()
                    ws = openpyxl.Workbook()
                    ws_sheet = ws.active
                    ws_sheet.append(['*SKU', '品名', '原FNSKU', '调整FNSKU', '调整量'])
                    sql = "select * from `amazon_form`.`pre_msku` where `SKU` = '%s' and `状态` = '未使用' and " \
                          "`国家` = '%s'" % (sku, country)
                    # print(sql)
                    self.cursor.execute(sql)
                    result = self.cursor.fetchall()
                    print(result)
                    if result and sku_num <= len(result):
                        k = 0
                        product_name, list_fnsku = self.get_productname(sku)
                        # print(list_fnsku)
                        print(len(list_fnsku))
                        for j in result:
                            if j['FNSKU'] in list_fnsku:
                                k += 1
                                self.change_sql(j['FNSKU'])
                                ws_sheet.append([sku, product_name, '', j['FNSKU']])
                                if k == sku_num:
                                    break
                            else:
                                print(j['FNSKU'])
                        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
                        self.connection.commit()
                        flag = self.check_sql(len(result), sku, country)
                        if flag:
                            self.mkdir()
                            ws.save(f'D:/换标调整/换标调整-{sku}-{time_now}.xlsx')
                            messagebox.showinfo(message='换标模板生成成功')
                            self.sku.set('')
                            self.num.set('')
                            os.startfile("D:/换标调整")
                        else:
                            messagebox.showinfo(message='FNSKU标记失败，请重试')
                    else:
                        messagebox.showinfo(message=f'{sku}的fnsku未使用个数不足')
                    self.sql_close()
            else:
                if not sku_num:
                    messagebox.showinfo(message='请输入要换标的FNSKU的个数')
                if not country:
                    messagebox.showinfo(message='请输入要换标的FNSKU的国家')
        else:
            if not sku and not filename:
                messagebox.showinfo(message='请先输入SKU和个数或者导入文件路径')

    def check_sql(self, length, sku, country):
        sql = "select * from `amazon_form`.`pre_msku` where `SKU` = '%s' and `状态` = '未使用' and " \
              "`国家` = '%s'" % (sku, country)
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        # print(len(result), length)
        if len(result) != length:
            return True
        else:
            return False

    def get_productname(self, sku):
        url = f"https://erp.lingxing.com/api/product/lists?search_field_time=create_time&sort_field=create_time&" \
              f"sort_type=desc&search_field=sku&search_value={sku}&attribute=&status=&is_matched_alibaba=&" \
              f"senior_search_list=[]&offset=0&is_combo=&length=500&is_aux=0&product_type[]=1&product_type[]=2&" \
              f"selected_product_ids=&req_time_sequence=%2Fapi%2Fproduct%2Flists$$"
        get_headers = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/erp/productManage'}
        get_msg = global_var.s.get(url, headers=get_headers)
        get_msg = json.loads(get_msg.text)
        productname = ''
        for i in get_msg['list']:
            if i['sku'] == sku:
                productname = i['product_name']
        auth_token = global_var.s2.cookies.get('auth-token')
        # print(auth_token)
        auth_token = auth_token.replace('%25', '%')
        auth_token = auth_token.replace('%23', '#')
        auth_token = auth_token.replace('%26', '&')
        auth_token = auth_token.replace('%2B', '+')
        auth_token = auth_token.replace('%28', '(')
        auth_token = auth_token.replace('%29', ')')
        auth_token = auth_token.replace('%2F', '/')
        auth_token = auth_token.replace('%3D', '=')
        auth_token = auth_token.replace('%3F', '?')
        post_url = 'https://gw.lingxingerp.com/listing-api/api/product/showOnline'
        post_headers = {'Host': 'gw.lingxingerp.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                        , 'Referer': 'https://erp.lingxing.com/',
                        'auth-token': auth_token,
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '909',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
        data = {"offset": 0, "length": 200, "search_field": "local_sku", "search_value": [f"{sku}"], "exact_search": 0,
                "sids": "", "status": "", "is_pair": "", "fulfillment_channel_type": "", "global_tag_ids": "",
                "req_time_sequence": "/listing-api/api/product/showOnline$$"}
        data_data = json.dumps(data)
        post_msg = global_var.s.post(post_url, headers=post_headers, data=data_data)
        post_msg = json.loads(post_msg.text)
        list_fnsku = []
        print(post_msg['data']['total'])
        print(len(post_msg['data']['list']))
        if post_msg['code'] == 1 and post_msg['msg'] == '成功':
            for i in post_msg['data']['list']:
                list_fnsku.append(i['fnsku'])
            if post_msg['data']['total'] > 200:
                data['offset'] = 200
                data_data = json.dumps(data)
                post_msg_other = global_var.s.post(post_url, headers=post_headers, data=data_data)
                post_msg_other = json.loads(post_msg_other.text)
                for i in post_msg_other['data']['list']:
                    list_fnsku.append(i['fnsku'])
                if post_msg['data']['total'] > 400:
                    data['offset'] = 400
                    data_data = json.dumps(data)
                    post_msg_other = global_var.s.post(post_url, headers=post_headers, data=data_data)
                    post_msg_other = json.loads(post_msg_other.text)
                    for i in post_msg_other['data']['list']:
                        list_fnsku.append(i['fnsku'])
                    if post_msg['data']['total'] > 600:
                        data['offset'] = 600
                        data_data = json.dumps(data)
                        post_msg_other = global_var.s.post(post_url, headers=post_headers, data=data_data)
                        post_msg_other = json.loads(post_msg_other.text)
                        for i in post_msg_other['data']['list']:
                            list_fnsku.append(i['fnsku'])
                        if post_msg['data']['total'] > 800:
                            data['offset'] = 800
                            data_data = json.dumps(data)
                            post_msg_other = global_var.s.post(post_url, headers=post_headers, data=data_data)
                            post_msg_other = json.loads(post_msg_other.text)
                            for i in post_msg_other['data']['list']:
                                list_fnsku.append(i['fnsku'])
        return productname, list_fnsku

    def get_inventory(self, sku):
        get_url = "https://erp.lingxing.com/api/storage/lists"
        get_headers = {'Host': 'erp.lingxing.com',
                       'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                       , 'Referer': 'https://erp.lingxing.com/erp/msupply/warehouseDetail',
                       'Accept': 'application/json, text/plain, */*',
                       'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                       'Accept-Encoding': 'gzip, deflate, br',
                       'Content-Type': 'application/json;charset=utf-8',
                       'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                       'X-AK-Company-Id': '90136229927150080',
                       'X-AK-Request-Source': 'erp',
                       'X-AK-ENV-KEY': 'SAAS-10',
                       'X-AK-Version': '1.0.0.0.0.023',
                       'X-AK-Zid': '109810',
                       'Content-Length': '909',
                       'Origin': 'https://erp.lingxing.com',
                       'Connection': 'keep-alive'}
        data = {"wid_list": "", "mid_list": "", "sid_list": "", "cid_list": "", "bid_list": "", "principal_list": "",
                "product_type_list": "", "product_attribute": "", "product_status": "", "search_field": "sku",
                "search_value": f"{sku}", "is_sku_merge_show": 0, "is_hide_zero_stock": 0, "offset": 0,
                "length": 200, "sort_field": "pre_total", "sort_type": "desc", "gtag_ids": "",
                "senior_search_list": "[]", "req_time_sequence": "/api/storage/lists$$"}
        data = json.dumps(data)
        get_msg = global_var.s.post(get_url, headers=get_headers, data=data)
        get_msg = json.loads(get_msg.text)
        num_prepare = 0
        print(get_msg)
        if get_msg['code'] == 1 and get_msg['msg'] == '操作成功':
            for i in get_msg['data']['list']:
                if i['sku'] == sku:
                    num_prepare += int(i['total'])
        return num_prepare

    def mkdir(self):
        folder = os.path.exists("D:/换标调整")
        if not folder:
            os.makedirs("D:/换标调整")

    def mkdir2(self, file):
        folder = os.path.exists(f"D:/换标调整/{file}")
        if not folder:
            os.makedirs(f"D:/换标调整/{file}")

    def get_msg(self, event=None):
        sku = self.sku.get().strip()
        country = self.country.get().strip()
        self.msg.set('')
        if sku and country:
            self.sql()
            sql = "select * from `amazon_form`.`pre_msku` where `SKU` = '%s' and `状态` = '未使用' and " \
                  "`国家` = '%s'" % (sku, country)
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            num = 0
            if result:
                num = len(result)
            self.sql_close()
            self.msg.set(f'{country}的{sku}有{num}个未使用的FNSKU')
        else:
            if not sku:
                messagebox.showinfo(message='请先输入要查询的SKU')
            if not country:
                messagebox.showinfo(message='请先输入要查询的国家')

    def read_excl(self):
        filename = self.itemname.get()
        warehouse = self.warehouse.get()
        if filename and warehouse:
            ask = messagebox.askokcancel(message=f'是否上传{filename}这个文件进行换标操作')
            if ask:
                try:
                    wb = openpyxl.load_workbook(filename)
                    wb_sheet = wb.active
                    row1 = wb_sheet.max_row
                    for i in range(row1, 0, -1):
                        cell_value1 = wb_sheet.cell(row=i, column=1).value
                        if cell_value1:
                            row1 = i
                            break
                    product_list = []
                    time_file = datetime.datetime.now().strftime("%Y%m%d")
                    self.mkdir2(f'fnsku_{time_file}')
                    # self.sql()
                    for i in range(2, row1+1):
                        sku = wb_sheet.cell(row=i, column=1).value.strip()
                        fnsku = wb_sheet.cell(row=i, column=3).value.strip()
                        fnsku_new = wb_sheet.cell(row=i, column=4).value.strip()
                        num = int(wb_sheet.cell(row=i, column=5).value)
                        num_inventory = self.get_inventory(sku)
                        print(num_inventory)
                        if num <= num_inventory:
                            data, msg = self.get_adjustmentSheet(sku, fnsku, fnsku_new, num, warehouse)
                            print(msg)
                            if data:
                                product_list.append(data)
                                self.get_fnsku_lable(sku, fnsku_new, f'fnsku_{time_file}')
                                # self.change_sql(fnsku_new)
                                wb_sheet.cell(row=i, column=6).value = '换标成功'
                            else:
                                wb_sheet.cell(row=i, column=6).value = '换标失败'
                        else:
                            wb_sheet.cell(row=i, column=6).value = '库存数量不足'
                            print('库存数量不足')
                    data = {}
                    data['af_id'] = 0
                    data['audit_users'] = []
                    data['data'] = []
                    data['order_sn'] = ""
                    data['data'] = product_list
                    data['remark'] = ""
                    data['req_time_sequence'] = '/api/storage/adjustment/order/submit$$'
                    data['type'] = 1
                    data['wid'] = self.get_warehouse(warehouse)
                    print(data)
                    data = json.dumps(data)
                    url = 'https://erp.lingxing.com/api/storage/adjustment/order/submit'
                    post_headers = {'Host': 'erp.lingxing.com',
                                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                                    , 'Referer': 'https://erp.lingxing.com/erp/msupply/adjustmentSheetAdd',
                                    'Accept': 'application/json, text/plain, */*',
                                    'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                                    'Accept-Encoding': 'gzip, deflate, br',
                                    'Content-Type': 'application/json;charset=utf-8',
                                    'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                                    'X-AK-Company-Id': '90136229927150080',
                                    'X-AK-Request-Source': 'erp',
                                    'X-AK-ENV-KEY': 'SAAS-10',
                                    'X-AK-Version': '1.0.0.0.0.023',
                                    'X-AK-Zid': '109810',
                                    'Content-Length': '909',
                                    'Origin': 'https://erp.lingxing.com',
                                    'Connection': 'keep-alive'}
                    post_msg = global_var.s.post(url, headers=post_headers, data=data)
                    post_msg = json.loads(post_msg.text)
                    print(post_msg)
                    if post_msg['code'] == 1 and post_msg['msg'] == '操作成功':
                        # self.connection.commit()
                        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
                        self.get_order_sn(post_msg['order_sn'], time_now)
                        # print_pdf.print_pdf(time_now, order_sn)
                        wb.save(f"D:/换标调整/{time_now}-{post_msg['order_sn']}/换标调整详情-{time_now}.xlsx")
                        shutil.move(f"D:/换标调整/fnsku_{time_file}/", rf"D:/换标调整/{time_now}-{post_msg['order_sn']}/")
                        messagebox.showinfo(message='换标调整完成')
                        os.startfile(f"D:/换标调整/{time_now}-{post_msg['order_sn']}")
                    else:
                        messagebox.showinfo(message=post_msg['msg'])
                    # self.sql_close()
                    self.itemname.set('')
                except Exception as e:
                    messagebox.showinfo(message=e)
        else:
            if not warehouse:
                messagebox.showinfo(message='请先选择仓库')
            if not filename:
                messagebox.showinfo(message='请先获取文件路径')

    def change_sql(self, fnsku):
        sql = "update `amazon_form`.`pre_msku` set `状态` = '已使用' where `FNSKU` = '%s'" % fnsku
        print(sql)
        self.cursor.execute(sql)
        self.connection.commit()

    def mkdir3(self, time_now, name):
        folder = os.path.exists(f"D:/换标调整/{time_now}-{name}")
        if not folder:
            os.makedirs(f"D:/换标调整/{time_now}-{name}")

    def get_adjustmentSheet(self, sku, fnsku, fnsku_new, num, warehouse):
        # print(fnsku)
        get_url = f"https://erp.lingxing.com/api/storage/adjustment/products/list?offset=0&length=200&key_value={sku}&product_type[]=1" \
                  f"&product_type[]=2&wid={self.get_warehouse(warehouse)}&req_time_sequence=/api/storage/fnskuLists$$"
        # print(get_url)
        get_headers = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/erp/msupply/adjustmentSheetAdd',
                       'Content-Type': 'application/json;charset=utf-8'}
        get_msg = global_var.s.get(get_url, headers=get_headers)
        get_msg = json.loads(get_msg.text)
        print(get_msg)
        data_fnsku = {}
        if get_msg['code'] == 1 and get_msg['msg'] == '操作成功' and get_msg['list']:
            for i in get_msg['list']:
                # print(i)
                if fnsku == i['fnsku']:
                    data_fnsku = i
                    # print('data-sku：',data_fnsku)
            if not data_fnsku:
                return False, 1
            data_fnsku['adjustment_valid_num'] = 0
            data_fnsku['adjustment_valid_sgn'] = '+'
            data_fnsku['adjustment_bad_num'] = 0
            data_fnsku['adjustment_bad_sgn'] = '+'
            data_fnsku['adjustment_available_bin'] = ""
            data_fnsku['adjustment_inferior_bin'] = ""
            data_fnsku['_XID'] = 'row_117'
            print(data_fnsku)
            data_fnsku_new = {}
            data_fnsku_new['data'] = []
            data_fnsku_new['data'].append({'fnsku': data_fnsku['fnsku'], 'product_id': data_fnsku['product_id'], 'seller_id':data_fnsku['seller_id']})
            data_fnsku_new['req_time_sequence'] = '/api/storage/adjustment/products/seller-fnsku$$'
            data_fnsku_new['wid'] = self.get_warehouse(warehouse)
            print(data_fnsku_new)
            # data.append(data_fnsku)
            data_fnsku_new = json.dumps(data_fnsku_new)
            post_url_fnsku = "https://erp.lingxing.com/api/storage/adjustment/products/seller-fnsku"
            post_headers_fnsku = {'Host': 'erp.lingxing.com',
                                  'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                                  , 'Referer': 'https://erp.lingxing.com/erp/msupply/adjustmentSheetAdd',
                                  'Accept': 'application/json, text/plain, */*',
                                  'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                                  'Accept-Encoding': 'gzip, deflate, br',
                                  'Content-Type': 'application/json;charset=utf-8',
                                  'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                                  'X-AK-Company-Id': '90136229927150080',
                                  'X-AK-Request-Source': 'erp',
                                  'X-AK-ENV-KEY': 'SAAS-10',
                                  'X-AK-Version': '1.0.0.0.0.023',
                                  'X-AK-Zid': '109810',
                                  'Content-Length': '909',
                                  'Origin': 'https://erp.lingxing.com',
                                  'Connection': 'keep-alive'}
            post_msg = global_var.s.post(post_url_fnsku, headers=post_headers_fnsku, data=data_fnsku_new)
            post_msg = json.loads(post_msg.text)
            print(post_msg)
            if post_msg['msg'] == '操作成功' and post_msg['code'] == 1 and post_msg['list']:
                for j in post_msg['list']:
                    if j['fnsku'] == fnsku:
                        # seller_name = {}
                        seller_name_new = {}
                        seller_id = ''
                        data_fnsku['sellerList'] = j['seller_list']
                        for i in j['seller_list']:
                            print (i['seller_name'])
                            if i['seller_name'] != "CoBak Direct-NA":
                                for k in i['fnsku_list']:
                                    if k['fnsku'] == fnsku_new:
                                        data_fnsku['fnskuList'] = i['fnsku_list']
                                        seller_id = i['seller_id']
                                        seller_name_new = i['seller_name']
                        # print(seller_name_new)
                        data_fnsku['adjustment_valid_num'] = num
                        data_fnsku['whb_in_list'] = []
                        data_fnsku['whb_out_list'] = []
                        data_fnsku['to_fnsku'] = fnsku_new
                        data_fnsku['to_product_valid_num'] = 0
                        data_fnsku['seller_id'] = str(j['seller_id'])
                        data_fnsku['to_seller_id'] = str(seller_id)
                        # data_fnsku['seller_name'] = seller_name
                        data_fnsku['to_seller_name'] = seller_name_new
                        return data_fnsku, True
                    else:
                        return False, 2
            else:
                return False, 3
        else:
            return False, 4

    def get_fnsku_lable(self, sku, fnsku, time_now):
        get_url = f"https://erp.lingxing.com/api/product/showOnline?start_date=2021-10-19&end_date=2021-10-19&" \
                  f"search_field=fnsku&search_value={fnsku}&status=&" \
                  "is_pair=&fulfillment_channel_type=&offset=0&length=50&req_time_sequence=/api/product/showOnline$$27"
        headers = {'Host': 'erp.lingxing.com',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0',
            # , 'Referer': 'https://erp.lingxing.com/erp/warehouse_detail',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
            'Accept-Encoding': 'gzip, deflate, br',
            # 'Content-Type': 'application/json;charset=utf-8',
            'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
            'X-AK-Company-Id': '90136229927150080',
            'X-AK-Request-Source': 'erp',
            'X-AK-ENV-KEY': 'SAAS-10',
            'X-AK-Version': '1.0.0.0.0.023',
            'X-AK-Zid': '109810',
            # 'Content-Length': '114',
            # 'Origin': 'https://erp.lingxing.com',
            'Connection': 'keep-alive'}
        please_get = global_var.s.get(get_url, headers=headers)
        please_get = json.loads(please_get.text)
        print(please_get)
        please_post = {}
        please_post['data'] = []
        please_post['data'].append(0)
        please_post['data'][0] = {}
        if please_get['msg'] and please_get['list']:
            for i in please_get['list']:
                if i['fnsku'] == fnsku and i['local_sku'] == sku:
                    data = {}
                    data['fnsku'] = fnsku
                    data['item_condition'] = i['item_condition']
                    data['item_name'] = i['item_name']
                    data['local_name'] = i['local_name']
                    data['local_sku'] = i['local_sku']
                    data['num'] = "1"
                    please_post['data'][0] = data
        please_post['is_content'] = 0
        please_post['is_self'] = 1
        please_post['name_type'] = 2
        please_post['page_type'] = 1
        please_post['req_time_sequence'] = "/api/print/printProduct$$8"
        please_post['self_content'] = "MADE IN CHINA"
        # print(please_post)
        please_post = json.dumps(please_post)
        print_url = "https://erp.lingxing.com/api/print/printProduct"
        post_headers = {'Host': 'erp.lingxing.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0',
                        'Referer': 'https://erp.lingxing.com/login',
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '114',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
        print_data = global_var.s.post(print_url, headers=post_headers, data=please_post)
        print_data = json.loads(print_data.text)
        # print(print_data)
        if print_data['msg'] == "操作成功":
            url2 = f"https://erp.lingxing.com/api/file/downloadById?id={int(print_data['file_id'])}&is_export=1"
            # 请求头
            headers = {'Host': 'erp.lingxing.com',
                       'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0',
                       # , 'Referer': 'https://erp.lingxing.com/erp/warehouse_detail',
                       'Accept': 'application/json, text/plain, */*',
                       'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                       'Accept-Encoding': 'gzip, deflate, br',
                       # 'Content-Type': 'application/json;charset=utf-8',
                       'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                       'X-AK-Company-Id': '90136229927150080',
                       'X-AK-Request-Source': 'erp',
                       'X-AK-ENV-KEY': 'SAAS-10',
                       'X-AK-Version': '1.0.0.0.0.023',
                       'X-AK-Zid': '109810',
                       # 'Content-Length': '114',
                       # 'Origin': 'https://erp.lingxing.com',
                       'Connection': 'keep-alive'}
            res2 = global_var.s.get(url2, headers=headers, stream=False)
            # print(res2)
            with open('D:/换标调整/%s/%s.pdf' % (time_now, fnsku), 'wb') as wr:
                wr.write(res2.content)

    def get_order_sn(self, msg_id, time_now):
        # order_sn = ''
        wb = openpyxl.Workbook()
        wb_sheet = wb.active
        wb_sheet.append(['单据编号', '仓库名称', '单据类型', '单据状态', '创建人', '创建时间', '操作人', '调整时间', 'SKU', '品名',
                         '店铺', 'FNSKU', '库存可用量', '调整数', '库存次品量', '调整数', '备注'])
        get_url = f"https://erp.lingxing.com/api/storage/adjustmentLists?search_field=order_sn&search_value={msg_id}&offset=0&length=20&" \
                  f"search_date_type=create_time&adjust_status=0&req_time_sequence=/api/storage/adjustmentLists$$"
        # print(get_url)
        get_headers = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/erp/msupply/adjustmentSheetAdd',
                       'Content-Type': 'application/json;charset=utf-8'}
        get_msg = global_var.s.get(get_url, headers=get_headers)
        get_msg = json.loads(get_msg.text)
        for i in get_msg['list']:
            if i['order_sn'] == msg_id:
                order_sn = i['order_sn']
                warehouse_name = i['ware_house_bak_name']
                type = i['type_text']
                status = i['status_text']
                realname = i['create_realname']
                creat_time = i['create_time']
                remark = i['remark']
                for j in i['item_list']:
                    wb_sheet.append([order_sn, warehouse_name, type, status, realname, creat_time, None, None,
                                     j['sku'], j['product_name'], j['seller_name'], j['fnsku'], j['product_valid_num'],
                                     j['adjustment_valid_num'], j['product_bad_num'], j['adjustment_bad_num'], remark])
        self.mkdir3(time_now, msg_id)
        wb.save(f'D:/换标调整/{time_now}-{msg_id}/换标调整单-{msg_id}.xlsx')
        # return order_sn

    def clear(self, event=None):
        ask = messagebox.askokcancel(message='是否清空当前界面信息')
        if ask:
            self.num.set('')
            self.sku.set('')
            self.country.set('')
            self.itemname.set('')
            self.warehouse.set('')
            self.msg.set('')
    # def get_excl(self):


class Paie_fnsku(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.root = master  # 定义内部变量root
        self.xscroll = Scrollbar(self, orient=HORIZONTAL)
        self.yscroll = Scrollbar(self, orient=VERTICAL)
        self.ft = tkFont.Font(family='microsoft yahei', size=10)
        self.creat()

    def creat(self):
        Label(self).grid(row=0, stick=W, pady=10)
        Button(self, text='配对', font=self.ft, command=self.pair_fnsku, width=30).grid(row=1, column=0, stick=W, pady=10, padx=10)

    def pair_fnsku(self):
        folder = os.path.exists("D:/listing")
        if not folder:
            os.makedirs("D:/listing")
        pair = find_msku.Find_order()
        pair.get_msku()
        messagebox.showinfo(message='配对完成')
