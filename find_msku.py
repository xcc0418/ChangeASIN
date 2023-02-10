import openpyxl
import requests
import pymysql
import json
import datetime
import zipfile
import time
import email.mime.multipart
import smtplib
from email.mime.text import MIMEText
# from email.mime.application import MIMEApplication
#
# import global_var


class Find_order():
    def __init__(self):
        self.s = requests.Session()
        login_url = 'https://erp.lingxing.com/api/passport/login'
        # 请求头
        headers = {'Host': 'erp.lingxing.com',
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                    , 'Referer': 'https://erp.lingxing.com/login',
                   'Accept': 'application/json, text/plain, */*',
                   'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                   'Accept-Encoding': 'gzip, deflate, br',
                   'Content-Type': 'application/json;charset=utf-8',
                   'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                   'X-AK-Company-Id': '90136229927150080',
                   'X-AK-Request-Source': 'erp',
                   'X-AK-ENV-KEY': 'SAAS-10',
                   'X-AK-Version': '1.0.0.0.0.023',
                   'X-AK-Zid': '109810',
                   'Content-Length': '114',
                   'Origin': 'https://erp.lingxing.com',
                   'Connection': 'keep-alive'}
        # 传递用户名和密码
        data = {'account': 'IT-Test', 'pwd': 'IT-Test'}
        data = json.dumps(data)
        self.s.post(login_url, headers=headers, data=data)
        self.auth_token = None
        self.cty_dict = {'美国': 1, '加拿大': 2, '墨西哥': 3, '英国': 4, '意大利': 7, '德国': 5, '法国': 6, '西班牙': 8,
                         '印度': 9, '日本': 10, '澳洲': 11, '阿联酋': 12, '新加坡': 13, '荷兰': 14, '沙特阿拉伯': 15,
                         '巴西': 16, '瑞典': 17, '土耳其': 20, '波兰': 19}
        # self.find_sql()

    def sql(self):
        self.connection = pymysql.connect(host='3354n8l084.goho.co',  # 数据库地址
                                          port=24824,
                                          user='test_user',  # 数据库用户名
                                          password='a123456',  # 数据库密码
                                          db='storage',  # 数据库名称
                                          charset='utf8',
                                          cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor()

    def sql_close(self):
        self.cursor.close()
        self.connection.close()

    def get_msku(self):
        # try:
            self.sql()
            self.delete_sql()
            sql = "select * from `amazon_form`.`pre_msku` where `状态` = '未匹配'"
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            if result:
                auth_token = self.s.cookies.get('auth-token')
                auth_token = auth_token.replace('%25', '%')
                auth_token = auth_token.replace('%23', '#')
                auth_token = auth_token.replace('%26', '&')
                auth_token = auth_token.replace('%2B', '+')
                auth_token = auth_token.replace('%28', '(')
                auth_token = auth_token.replace('%29', ')')
                auth_token = auth_token.replace('%2F', '/')
                auth_token = auth_token.replace('%3D', '=')
                auth_token = auth_token.replace('%3F', '?')
                self.auth_token = auth_token
                list_msku = []
                    # print(auth_token)
                for i in result:
                    msku = i['MSKU'].strip()
                    sku = i['SKU'].strip()
                    # asin = i['ASIN']
                    country = i['国家'].strip()
                    supplier = i['供应商'].strip()
                    fnsku, asin, shop_name = self.find_fnsku(msku, country, auth_token)
                    print(asin)
                    if fnsku:
                        product_name = self.get_productname(sku)
                        # list_asin = self.get_asin(sku, auth_token)
                        list_msku.append([asin, country, fnsku, supplier, sku, 1.0, shop_name, msku, product_name])
                print(list_msku)
                index, list_asin = self.downloads(auth_token)
                self.sql_close()
                if list_asin:
                    self.sql()
                    for i in list_msku:
                        if list_asin.count(i[0]) > 1:
                            sql = "update `amazon_form`.`pre_msku` set `FNSKU` = '%s' , `状态` = '存在相同ASIN', `ASIN` = '%s' " \
                                  "where `MSKU` = '%s'" % (i[2], i[0], i[7])
                            # print(111)
                        else:
                            print(i)
                            self.peidui(i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7], i[8])
                            sql = "update `amazon_form`.`pre_msku` set `FNSKU` = '%s' , `状态` = '未使用', `ASIN` = '%s' " \
                                  "where `MSKU` = '%s'" % (i[2], i[0], i[7])
                            list_asin.append(i[0])
                        self.cursor.execute(sql)
                    self.connection.commit()
                else:
                    self.smtplib_error(index)
            self.sql_close()
        # except Exception as e:
        #     print(e)

    def get_productname(self, sku):
        url = f"https://erp.lingxing.com/api/product/lists?search_field_time=create_time&sort_field=create_time&" \
              f"sort_type=desc&search_field=sku&search_value={sku}&attribute=&status=&is_matched_alibaba=&" \
              f"senior_search_list=[]&offset=0&is_combo=&length=20&is_aux=0&product_type[]=1&product_type[]=2&" \
              f"selected_product_ids=&req_time_sequence=%2Fapi%2Fproduct%2Flists$$"
        get_headers = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/erp/productManage'}
        get_msg = self.s.get(url, headers=get_headers)
        get_msg = json.loads(get_msg.text)
        productname = ''
        for i in get_msg['list']:
            if i['sku'] == sku:
                productname = i['product_name']
        return productname

    # def get_asin(self, sku, auth_token):
    #     post_url = "https://gw.lingxingerp.com/listing-api/api/product/showOnline"
    #     post_headers = {'Host': 'gw.lingxingerp.com',
    #                     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
    #                     , 'Referer': 'https://erp.lingxing.com',
    #                     'auth-token': auth_token,
    #                     'Accept': 'application/json, text/plain, */*',
    #                     'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
    #                     'Accept-Encoding': 'gzip, deflate, br',
    #                     'Content-Type': 'application/json;charset=utf-8',
    #                     'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
    #                     'X-AK-Company-Id': '90136229927150080',
    #                     'X-AK-Request-Source': 'erp',
    #                     'X-AK-ENV-KEY': 'SAAS-10',
    #                     'X-AK-Version': '1.0.0.0.0.023',
    #                     'X-AK-Zid': '109810',
    #                     'Content-Length': '909',
    #                     'Origin': 'https://erp.lingxing.com',
    #                     'Connection': 'keep-alive'}
    #     data = {}
    #     data['fulfillment_channel_type'] = ''
    #     data['is_pair'] = 1
    #     data['length'] = 200
    #     data['offset'] = 0
    #     data['req_time_sequence'] = '/listing-api/api/product/showOnline$$'
    #     data['search_field'] = 'local_sku'
    #     data['search_value'] = []
    #     data['search_value'].append(sku)
    #     data['sids'] = ''
    #     data['status'] = ''
    #     # print(data)
    #     data = json.dumps(data)
    #     post_msg = self.s.post(post_url, headers=post_headers, data=data)
    #     post_msg = json.loads(post_msg.text)
    #     list_asin = []
    #     if post_msg['code'] == 1 and post_msg['msg'] == '成功' and post_msg['data']['list']:
    #         for i in post_msg['data']['list']:
    #             list_asin.append(i['asin'])
    #     return list_asin

    def downloads(self, auth_token):
        try:
            url = "https://gw.lingxingerp.com/listing-api/api/product/exportOnline"
            headers = {'Host': 'gw.lingxingerp.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                        , 'Referer': 'https://erp.lingxing.com',
                       'auth-token': auth_token,
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '909',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
            data = {"offset": 0, "length": 50, "search_field": "local_sku", "exact_search": 1, "sids": "", "status": "",
                    "is_pair": "", "fulfillment_channel_type": "", "global_tag_ids": "",
                    "req_time_sequence": "/listing-api/api/product/exportOnline$$2"}
            data = json.dumps(data)
            post_msg = self.s.post(url, headers=headers, data=data)
            post_msg = json.loads(post_msg.text)
            print(post_msg)
            if post_msg['code'] == 1 and post_msg['msg'] == "成功":
                report_id = post_msg['data']['data']['report_id']
                time.sleep(100)
                if report_id:
                    file_download_url = f"https://erp.lingxing.com/api/download/downloadCenterReport/downloadResource?report_id={report_id}"
                    # print(file_download_url)
                    get_headers = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/erp/muser/downloadCenter'}
                    download_file = self.s.get(file_download_url, headers=get_headers, stream=False)
                    with open('D:/listing/listing.zip', 'wb') as q:
                        q.write(download_file.content)
                    list_asin = self.read_excl(report_id)
                    if list_asin:
                        return True, list_asin
                    else:
                        return False, False
                else:
                    return False, False
            else:
                return False, False
        except Exception as e:
            print(e)
            return False, e

    def read_excl(self, report):
        file = zipfile.ZipFile('D:/listing/listing.zip')
        file.extractall('D:/listing/')
        file.close()
        time_data = datetime.datetime.now().strftime("%Y%m%d")
        filename = f"D:/listing/listing{time_data}-{report}.xlsx"
        wb = openpyxl.load_workbook(filename)
        wb_sheet = wb.active
        row_max = wb_sheet.max_row
        list_asin = []
        for i in range(2, row_max+1):
            asin = wb_sheet.cell(row=i, column=5).value
            if asin:
                if asin in list_asin or len(asin) != 10 or asin.find('B0') < 0:
                    continue
                else:
                    list_asin.append(asin)
        if len(list_asin) > 5000 and 'B0BLCS9J5G' in list_asin:
            return list_asin
        else:
            return False

    def find_fnsku(self, msku, country, auth_token):
        post_url = "https://gw.lingxingerp.com/listing-api/api/product/showOnline"
        post_headers = {'Host': 'gw.lingxingerp.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                        , 'Referer': 'https://erp.lingxing.com',
                        'auth-token': auth_token,
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '909',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
        data = {}
        data['fulfillment_channel_type'] = ''
        data['is_pair'] = ''
        data['length'] = 200
        data['offset'] = 0
        data['req_time_sequence'] = '/listing-api/api/product/showOnline$$'
        data['search_field'] = 'msku'
        data['search_value'] = []
        data['search_value'].append(msku)
        data['sids'] = ''
        data['status'] = ''
        # print(data)
        data = json.dumps(data)
        post_msg = self.s.post(post_url, headers=post_headers, data=data)
        post_msg = json.loads(post_msg.text)
        # print(post_msg)
        # print(msku)
        if post_msg['code'] == 1 and post_msg['msg'] == '成功' and post_msg['data']['list']:
            for i in post_msg['data']['list']:
                # msku_ = i['msku'].strip()
                if i['msku'] == msku and country == i['marketplace']:
                    # print(i['fnsku'])
                    return i['fnsku'], i['asin'], i['seller_name']
                else:
                    print(i['msku'])
            return False, False, False
        return False, False, False

    def peidui(self, asin, country, fnsku, supplier, sku, version, shop_name, msku, product_name):
        key = country + fnsku
        sql3 = f"INSERT INTO data_read.listing VALUES('{key}','{shop_name}','{country}'," \
               f"'{asin}','{msku}','{fnsku}','{product_name}','{sku}','可采购'," \
               f"'可发货', '{version}',85,'60','{supplier}',DEFAULT,DEFAULT)"
        try:
            self.cursor.execute(sql3)
            file_url2 = f"https://erp.lingxing.com/api/product/lists?search_field=sku&search_value={sku}"
            # 请求头
            headers2 = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/productManage'}
            res2 = self.s.get(file_url2, headers=headers2, stream=False)
            res2 = res2.text
            res2 = json.loads(res2)
            id = None
            for i in res2['list']:
                if sku == i['sku']:
                    id = i['id']
            url3 = "https://gw.lingxingerp.com/listing-api/api/product/batchLink"
            url4 = "https://gw.lingxingerp.com/listing-api/api/product/showOnline"
            # print(auth_token)
            headers3 = {
                        'Host': 'gw.lingxingerp.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:101.0) Gecko/20100101 Firefox/101.0'
                        , 'Referer': 'https://erp.lingxing.com/',
                        'Accept': 'application/json, text/plain, */*',
                        'auth-token': self.auth_token,
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '2.9.5.0.1.010',
                        'X-AK-Zid': '109810',
                        'Content-Length': '164',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
            data3 = {"sid_msku": [], "pid": id, "is_sync_pic": 0,
                     "req_time_sequence": "/listing-api/api/product/batchLink$$"}
            data4 = {}
            data4['fulfillment_channel_type'] = ''
            data4['is_pair'] = ''
            data4['req_time_sequence'] = '/listing-api/api/product/showOnline$$'
            data4['search_field'] = 'asin'
            data4['search_value'] = [asin]
            data4['status'] = ''
            data4 = json.dumps(data4)
            res4 = self.s.post(url4, headers=headers3, data=data4)
            res4 = json.loads(res4.text)
            # print(res4)
            if res4['code'] != 1:
                raise Exception('配对失败。')
            else:
                for i in res4['data']['list']:
                    if i['fnsku'] == fnsku:
                        data3['sid_msku'].append({'msku': i['msku'], 'store_id': i['store_id']})
            data3 = json.dumps(data3)
            res3 = self.s.post(url3, headers=headers3, data=data3)
            a3 = res3.text
            b3 = json.loads(a3)
            print(b3)
            if b3['code'] != 1:
                raise Exception('ERP中没配对成功。')
            self.connection.commit()
        except Exception as e:
            self.connection.rollback()
            print(e)
            print("配对失败！！！")
            return False, "配对失败！！！"
        return 2, b3['msg']

    def max_num(self, a, b):
        if float(a) > float(b):
            return float(a)
        else:
            return float(b)

    def delete_sql(self):
        time_now = datetime.datetime.now().strftime("%Y-%m-%d")
        time_now = datetime.datetime.strptime(time_now, '%Y-%m-%d')
        sql = "select * from `amazon_form`.`pre_msku` where `状态` = '未匹配'"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        # print(result)
        for i in result:
            if i['ASIN'] and i['FNSKU']:
                continue
            else:
                time_data = str(i['创建时间'])
                time_data = time_data[0:10]
                # print(time_data)
                time_data = datetime.datetime.strptime(time_data, '%Y-%m-%d')
                # time_data = datetime.datetime.strptime(time_data, '%Y-%m-%d-%H-%M-%S')
                # time_data = time_data[0:10]
                # print(time_data, time_now)
                data_day = time_now - time_data
                # print(data_day.days)
                if data_day.days >= 7:
                    msku = i['MSKU']
                    sql1 = "update `amazon_form`.`pre_msku` set `状态` = '已删除'" \
                           "where `MSKU` = '%s'" % msku
                    self.cursor.execute(sql1)
        self.connection.commit()

    def smtplib_error(self, e):
        time_now = time.strftime("%Y-%m-%d", time.localtime())
        my_send = 'xiechangcong@getoo.store'
        my_password = 'XCc200418'
        receivers = '2371138547@qq.com'  # 接收邮件，可设置为你的QQ邮箱或者其他邮箱 nixuemin@getoo.store  2371138547@qq.com
        ret = True
        # 三个参数：第一个为文本内容，第二个 plain 设置文本格式，第三个 utf-8 设置编码
        message = email.mime.multipart.MIMEMultipart()
        message['from'] = my_send
        message['to'] = receivers
        subject = 'MSKU配对'
        message['Subject'] = subject
        content = f"{e}"
        text = email.mime.text.MIMEText(content, 'plain', 'utf-8')
        message.attach(text)
        # att = MIMEApplication (open (f"{filename}", 'rb').read ())
        # att.add_header ('Content-Disposition', 'attachment', filename=f"常用物料采购-{time_now}.xlsx")
        # message.attach (att)
        try:
            # message['From'] = formataddr(["XCC", my_send])  # 发送者
            # message['To'] = formataddr(["测试", receivers])  # 接收者
            server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)
            server.login(my_send, my_password)
            server.sendmail(my_send, [receivers, ], message.as_string())
            server.quit()
        except Exception as e:
            print(e)
            ret = False
        if ret:
            print("邮件发送成功")
        else:
            print("邮件发送失败")


if __name__ == '__main__':
    find = Find_order()
    find.get_msku()
